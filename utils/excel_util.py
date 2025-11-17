'''
Excel操作模块
'''
import string
import openpyxl
import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), './')))



def create_sheet(file_name, sheet_name):
    """excel新增sheet
    Args:
        file_name (path): excel文件名   *.xlsx
        sheet_name (str): sheet名称
    Returns:
        sheet: excel sheet对象
    """
    if not os.path.isfile(file_name):
        # 创建一个新的工作簿
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(filename=file_name)
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(title=sheet_name)
    workbook.save(filename=file_name)
    return workbook[sheet_name]

def remove_sheet(file_name, sheet_name='Sheet'):
    if not os.path.isfile(file_name):
        # 创建一个新的工作簿
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(filename=file_name)
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    workbook.save(filename=file_name)


def write_line(file_name, sheet_name, row=0, line=None):
    """excel写入一行数据
    Args:
        file_name (path): excel文件名
        sheet_name(str): sheet名
        row (int): 行号
        line (list[dict]): 行内容,eg: [{"content": "IMEI", "style": STYLE}, {"content": "861881056520894"}]
    """
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet = workbook[sheet_name]
    if not row:
        row = sheet.max_row+1
    if not line:
        line = list()
    for i, cell_content in enumerate(line):
        cell = sheet.cell(row=row, column=i+1,
                          value=cell_content.get("content"))
        if cell_content.get("style"):
            cell.style = cell_content.get("style") 
    workbook.save(filename=file_name)


def merge_cells(file_name, sheet_name, from_row=0, from_column=1, row_count=1, column_count=1):
    """合并单元格
    Args:
        file_name (_type_): _description_
        sheet_name (_type_): _description_
        from_row (_type_): _description_
        from_column (_type_): _description_
        row_count (int, optional): _description_. Defaults to 1.
        column_count (int, optional): _description_. Defaults to 1.
    """
    workbook = openpyxl.load_workbook(filename=file_name)
    sheet = workbook[sheet_name]
    if not from_row:
        from_row = sheet.max_row + 1
    from_cell = f"{get_column_range_end(from_column)}{from_row}"
    end_cell = f"{get_column_range_end(from_column+column_count-1)}{from_row+row_count-1}"
    sheet.merge_cells(f"{from_cell}:{end_cell}")
    workbook.save(filename=file_name)


def get_column_range_end(column_count):
    """根据列索引，生成最后一列的字母名
    Args:
        column_count (_type_): 
    Returns:
        _type_: 
    """
    ret = ''
    ci = column_count - 1
    index = ci // 26
    if index > 0:
        ret += get_column_range_end(index)
    ret += string.ascii_uppercase[ci % 26]
    return ret
